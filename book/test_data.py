# -*- coding: UTF-8 -*-
import os

import openpyxl
import xlrd


def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    # 获取所有工作表名
    names = wb.sheetnames
    # wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
    sheet = wb[names[0]]
    # 获取最大行数
    maxRow = sheet.max_row
    # 获取最大列数
    maxColumn = sheet.max_column
    # 获取当前活动表
    current_sheet = wb.active
    # 获取当前活动表名称
    current_name = sheet.title
    # 通过名字访问Cell对象, 通过value属性获取值
    a1 = sheet['A1'].value
    # 通过行和列确定数据
    a12 = sheet.cell(row=1, column=2).value
    # 获取列字母
    column_name = openpyxl.utils.cell.get_column_letter(1)
    # 将列字母转为数字, 参数忽略大小写
    column_name_num = openpyxl.utils.cell.column_index_from_string('a')
    # 获取一列数据, sheet.iter_rows() 获取所有的行
    """
    (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>)
    (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>)
    (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>)
    (<Cell 'Sheet1'.A4>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.C4>)
    (<Cell 'Sheet1'.A5>, <Cell 'Sheet1'.B5>, <Cell 'Sheet1'.C5>)
    """
    datas = []
    for one_column_data in sheet.iter_rows():
        items = []
        items.append(one_column_data[0].value)
        items.append(one_column_data[1].value)
        items.append(one_column_data[2].value)
        datas.append(items)
    print(datas)
    return datas
    # 获取一行数据, sheet.iter_cols() 获取所有的列
    """
    (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>)
    (<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>)
    (<Cell 'Sheet1'.C1>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.C3>)
    """
    # for one_row_data in sheet.iter_cols():
    #     print(one_row_data[0].value)

def process_test_data():
    # 打开文件
    data = xlrd.open_workbook('E:\\BOOK_DATA\\测试集\\所有教材_书高.xls')
    # 查看工作表
    data.sheet_names()
    print("sheets：" + str(data.sheet_names()))
    # 读取sheet
    sheet = data.sheet_by_index(0)

    # 读取文件名和bookid的对应关系
    name_dict = {}
    for row_num in range(sheet.nrows):
        try:
            row_value = sheet.row_values(row_num)
            book_id = int(row_value[0])
            url = row_value[2]
            url = str(url).replace("https://universe-file.tuling123.com/book_image/", "").replace(".jpg", "")
            name_dict[url] = book_id
        except Exception as e:
            print(e)
    print(name_dict)
    # 重命名所有文件
    path = 'E:\\BOOK_DATA\\测试集\\所有教材原始图片\\'
    files = os.listdir(path)
    for file in files:
        print(file)
        file_code = file.replace(".jpg", "")
        old_name = path + file
        new_name = path + str(name_dict[file_code]) + ".jpg"
        print(old_name, new_name)
        os.renames(old_name, new_name)


if __name__ == '__main__':
    read_excel('E:\\BOOK_DATA\\测试集\\所有教材_书高.xlsx')
